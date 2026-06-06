import re
import uuid
from pathlib import Path
from app.config import settings
from app.chunker import split_text_with_images


def process_document(file_path: str) -> list[dict]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = path.suffix.lower()
    doc_id = str(uuid.uuid4())
    source = path.name

    if suffix == ".txt":
        return _process_txt(path, doc_id, source)
    elif suffix == ".md":
        return _process_markdown(path, doc_id, source)
    elif suffix == ".pdf":
        return _process_pdf(path, doc_id, source)
    elif suffix == ".xlsx" or suffix == ".xls":
        return _process_xlsx(path, doc_id, source)
    elif suffix == ".docx":
        return _process_docx(path, doc_id, source)
    else:
        raise ValueError(f"Unsupported file format: {suffix}")


def _process_txt(path: Path, doc_id: str, source: str) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="replace")
    chunks = split_text_with_images(text, [])
    for c in chunks:
        c["source"] = source
        c["doc_id"] = doc_id
        c["doc_type"] = "txt"
    return chunks


def _process_markdown(path: Path, doc_id: str, source: str) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="replace")
    image_positions = []
    img_pattern = re.compile(r"!\[([^\]]*)\]\(([^)]+)\)")
    for m in img_pattern.finditer(text):
        img_path = m.group(2)
        resolved = _resolve_image_path(img_path, path.parent)
        if resolved:
            image_positions.append({"path": str(resolved), "char_position": m.start()})
    clean_text = img_pattern.sub("", text)
    chunks = split_text_with_images(clean_text, image_positions)
    for c in chunks:
        c["source"] = source
        c["doc_id"] = doc_id
        c["doc_type"] = "markdown"
    return chunks


def _process_pdf(path: Path, doc_id: str, source: str) -> list[dict]:
    import fitz
    doc = fitz.open(path)
    all_chunks = []
    current_pos = 0

    for page_num, page in enumerate(doc):
        page_text = page.get_text()
        image_list = page.get_images(full=True)
        image_positions = []

        for img_idx, img in enumerate(image_list):
            xref = img[0]
            base_image = doc.extract_image(xref)
            img_bytes = base_image["image"]
            img_ext = base_image["ext"]
            img_filename = f"{doc_id}_p{page_num + 1}_img{img_idx}.{img_ext}"
            img_path = settings.image_dir / img_filename
            img_path.write_bytes(img_bytes)
            image_positions.append({"path": str(img_path), "char_position": current_pos})

        page_chunks = split_text_with_images(page_text, image_positions)
        for c in page_chunks:
            c["source"] = f"{source} (p.{page_num + 1})"
            c["doc_id"] = doc_id
            c["doc_type"] = "pdf"
            c["page"] = page_num + 1
        all_chunks.extend(page_chunks)
        current_pos += len(page_text)

    doc.close()
    return all_chunks


def _process_docx(path: Path, doc_id: str, source: str) -> list[dict]:
    from docx import Document
    import uuid as _uuid

    doc = Document(path)
    elements = []

    for para in doc.paragraphs:
        if para.text.strip():
            elements.append({"type": "text", "content": para.text.strip()})

    for rel in doc.part.rels.values():
        if "image" in rel.reltype:
            try:
                image_data = rel.target_part.blob
                ext = rel.target_ref.rsplit(".", 1)[-1] if "." in rel.target_ref else "png"
                img_filename = f"{doc_id}_{_uuid.uuid4().hex[:8]}.{ext}"
                img_path = settings.image_dir / img_filename
                img_path.write_bytes(image_data)
                elements.append({"type": "image", "path": str(img_path)})
            except Exception:
                pass

    full_text = ""
    image_positions = []
    char_pos = 0
    for elem in elements:
        if elem["type"] == "text":
            full_text += elem["content"] + "\n\n"
            char_pos += len(elem["content"]) + 2
        elif elem["type"] == "image":
            image_positions.append({"path": elem["path"], "char_position": char_pos})

    chunks = split_text_with_images(full_text, image_positions)
    for c in chunks:
        c["source"] = source
        c["doc_id"] = doc_id
        c["doc_type"] = "docx"
    return chunks


def _resolve_image_path(img_path: str, base_dir: Path):
    img_path = img_path.split()[0]
    candidate = Path(img_path)
    if candidate.is_file():
        return candidate
    candidate = base_dir / img_path
    if candidate.is_file():
        return candidate
    candidate = settings.upload_dir / img_path
    if candidate.is_file():
        return candidate
    return None

def _process_xlsx(path: Path, doc_id: str, source: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    full_text = ""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            line = "\t".join(cells)
            if line.strip():
                rows.append(line)
        if rows:
            full_text += f"[Sheet: {sheet_name}]\n" + "\n".join(rows) + "\n\n"
    wb.close()

    chunks = split_text_with_images(full_text, [])
    for c in chunks:
        c["source"] = source
        c["doc_id"] = doc_id
        c["doc_type"] = "xlsx"
    return chunks
